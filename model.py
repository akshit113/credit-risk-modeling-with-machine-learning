# -*- coding: utf-8 -*-
"""

__author__ = 'Akshit Agarwal'
__email__ = 'akshit@email.arizona.edu'
__date__ = '2020-11-29'
__dataset__ = 'https://www.kaggle.com/c/GiveMeSomeCredit/'
__connect__ = 'https://www.linkedin.com/in/akshit-agarwal93/'

"""

import glob
import warnings
from datetime import datetime

import numpy as np
from pandas import read_csv, cut, DataFrame, concat, read_excel, get_dummies
from sklearn.metrics import f1_score, jaccard_score, precision_score, recall_score, roc_auc_score
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
from sklearn.utils import resample
from tensorflow.keras.layers import Dropout, Dense
from tensorflow.python.keras.models import Sequential


def fxn():
    warnings.warn("deprecated", DeprecationWarning)


with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    fxn()

import sys
import numpy

numpy.set_printoptions(threshold=sys.maxsize)


def import_data(fname, train=True):
    """Import dataset and remove row numbers column
    :return: dataframe
    """
    df = read_csv(fname)
    if train:
        df = df.iloc[:, [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 1]]
    else:
        df = df.iloc[:, [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]]
    return df


def clean_data(df):
    """Clean dataframe and impute missing values by mean and mode
    :param df: input dataframe
    :return: clean dataframe
    """
    df['ages'] = cut(x=df['age'], bins=[0, 9, 19, 29, 39, 49, 59, 69, 79, 89, 99, 159],
                     labels=['0s', '10s', '20s', '30s', '40s', '50s', '60s', '70s', '80s', '90s', '100s'])
    mode_age = int(df['NumberOfDependents'].mode())
    df['NumberOfDependents'] = df['NumberOfDependents'].fillna(mode_age)
    mean_income = int(df['MonthlyIncome'].mean())
    df['MonthlyIncome'] = df['MonthlyIncome'].fillna(mean_income)
    df.dropna(inplace=True)
    missing = (df.isnull().values.any())
    if missing:
        print(df.isnull().sum())
    return df


def normalize_columns(df, colnames):
    """Performs Normalization using MinMaxScaler class in Sckikit-learn"""
    scaler = MinMaxScaler()
    for col in colnames:
        # Create x, where x the 'scores' column's values as floats
        x = df.loc[:, [col]]
        x = df[[col]].values.astype(float)
        # Create a minimum and maximum processor object
        # Create an object to transform the data to fit minmax processor
        x_scaled = scaler.fit_transform(x)
        # Run the normalizer on the dataframe
        df[col] = DataFrame(x_scaled)
    print(f'''Normalized Columns: {colnames} using MinMaxScaler.''')
    return df


def one_hot_encode(df, test_df, colnames):
    """This function performs one-hot encoding of the columns
    :param df: input df
    :param colnames: columns to be one-hot encoded
    :return: dataframe
    """

    for col in colnames:
        oh_df = get_dummies(df[col], prefix=col, drop_first=True)
        df = concat([oh_df, df], axis=1)
        df = df.drop([col], axis=1)
    missing = (df.isnull().values.any())
    while missing:
        df = df.dropna()
        print(df.isnull().sum())
        missing = (df.isnull().values.any())

    print(df.shape)
    print(list(df.columns))
    print(df.shape)
    return df


def undersample(df):
    # Separate majority and minority classes
    df_majority = df[df['SeriousDlqin2yrs'] == 0]
    df_minority = df[df['SeriousDlqin2yrs'] == 1]
    samples = min(len(df_minority), len(df_majority))
    # Downsample majority class
    df_majority_downsampled = resample(df_majority,
                                       replace=False,  # sample without replacement
                                       n_samples=samples,  # to match minority class
                                       random_state=42)  # reproducible results

    # Combine minority class with downsampled majority class
    df_downsampled = concat([df_majority_downsampled, df_minority])

    # Display new class counts
    print(df_downsampled['SeriousDlqin2yrs'].value_counts())
    # 1    49
    # 0    49
    # Name: balance, dtype: int64
    return df_downsampled


def print_cols(df):
    for idx, val in enumerate(df.columns, start=1):
        print(idx, ' ', val)


def reorder_cols(df):
    df = df.iloc[:, []]


def split_dataset(df, test_size, seed):
    """This function randomly splits (using seed) train data into training set and validation set. The test size
    paramter specifies the ratio of input that must be allocated to the test set
    :param df: one-hot encoded dataset
    :param test_size: ratio of test-train data
    :param seed: random split
    :return: training and validation data
    """
    ncols = np.size(df, 1)
    X = df.iloc[:, range(0, ncols - 1)]
    Y = df.iloc[:, ncols - 1]
    x_train, x_test, y_train, y_test = train_test_split(X, Y, test_size=test_size, random_state=seed)
    print(x_train.shape)
    print(x_test.shape)
    print(y_train.shape)
    print(y_test.shape)
    return X, Y, x_train, x_test, y_train, y_test


def get_model(input_size, output_size, magic='relu'):
    """This function creates a baseline feedforward neural network with of given input size and output size
        using magic activation function.
    :param input_size: number of columns in x_train
    :param output_size: no of columns in one hpt
    :param magic: activation function
    :return:Sequential model
    """
    dropout_rate = 0.2
    mlmodel = Sequential()
    mlmodel.add(Dense(18, input_dim=input_size, activation='selu'))

    mlmodel.add(Dense(32, activation='selu', kernel_initializer="uniform"))
    mlmodel.add(Dropout(dropout_rate))
    mlmodel.add(Dense(32, activation='selu', kernel_initializer="uniform"))
    mlmodel.add(Dropout(dropout_rate))
    mlmodel.add(Dense(128, activation='selu', kernel_initializer="uniform"))
    mlmodel.add(Dropout(dropout_rate))
    mlmodel.add(Dense(128, activation='selu', kernel_initializer="uniform"))
    mlmodel.add(Dropout(dropout_rate))
    mlmodel.add(Dense(256, activation='selu', kernel_initializer="uniform"))
    mlmodel.add(Dropout(dropout_rate))
    mlmodel.add(Dense(256, activation='selu', kernel_initializer="uniform"))
    mlmodel.add(Dropout(dropout_rate))
    mlmodel.add(Dense(512, activation='selu', kernel_initializer="uniform"))
    mlmodel.add(Dropout(dropout_rate))
    mlmodel.add(Dense(512, activation='selu', kernel_initializer="uniform"))
    mlmodel.add(Dropout(dropout_rate))
    mlmodel.add(Dense(1024, activation='selu', kernel_initializer="uniform"))
    mlmodel.add(Dropout(dropout_rate))
    mlmodel.add(Dense(1024, activation='selu', kernel_initializer="uniform"))

    mlmodel.add(Dense(1, activation='sigmoid', kernel_initializer="uniform"))

    # Setting optimizer
    # mlmodel.compile(loss="binary_crossentropy", optimizer='adam', metrics=['accuracy'])
    mlmodel.compile(optimizer='adam', loss='binary_crossentropy', metrics=['accuracy'])
    # opt = SGD(lr=0.01)
    # mlmodel.compile(loss="binary_crossentropy", optimizer=opt, metrics=['binary_accuracy'])
    return mlmodel


def fit_and_evaluate(model, x_train, y_train, x_test, y_test, batch_size, epochs):
    """fits the model created in the get_model function on x_train, y_train and evaluates the model performance on
    x_test and y_test using the batch size and epochs paramters
    :param model: Sequential model
    :param x_train: training data
    :param y_train: training label
    :param x_test: testing data
    :param y_test: testing label
    :param batch_size: amount of training data (x_train) fed to the model
    :param epochs: number of times the entire dataset is passed through the network
    :return: tuple of validation_accuracy and validation_loss
    """

    model.fit(x_train, y_train, batch_size=batch_size, epochs=epochs, validation_data=(x_test, y_test))
    test_loss, test_acc = model.evaluate(x_test, y_test, verbose=2)
    test_acc = round(test_acc * 100, 2)
    test_loss = round(test_loss * 100, 2)
    print('Test accuracy:', test_acc)
    print('Test Loss:', test_loss)
    return test_acc, test_loss


def get_metrics(y, y_hat, metric, average):
    """This function evaluates the model predictions, y_hat with ground truth y using a sklearn metric
    :param y: ground truth
    :param y_hat: model predictions
    :param metric: evaluation metric (f1_score, precision_score, jaccard_score
    :param average: micro, macro, binary, weighted
    :return: evaluation score
    """
    score = 0
    metrics_list = ['f1_score', 'jaccard_score', 'precision_score', 'recall_score', 'roc_auc_score']
    average_list = ['micro', 'macro', 'binary']
    if metric not in metrics_list:
        print(f'''{metric} is not a valid metric type. Please try one of these: {metrics_list}''')
        return
    if average not in average_list:
        print(f'''{average} is not a valid average type. Please try one of these: {average_list}''')
        return
    if metric == 'f1_score':
        score = f1_score(y, y_hat, average=average)
    elif metric == 'jaccard_score':
        score = jaccard_score(y, y_hat, average=average)
    elif metric == 'precision_score':
        score = precision_score(y, y_hat, average=average)
    elif metric == 'recall_score':
        score = recall_score(y, y_hat, average=average)
    elif metric == 'roc_auc_score':
        score = roc_auc_score(y, y_hat, average=average)
    score = round(score, 4)
    print(f'''{metric}: {score}''')
    return score


def make_predictions(model, x_test):
    """This function makes predictions using the model on the unseen test dataset
    :param y_test: test labels
    :param model: Sequential model
    :param x_test: unseen test dataset
    :return: predictions in the binary numpy array format
    """
    print(x_test)
    print(x_test.shape)
    print(list(x_test.columns))
    predictions = model.predict(x_test)
    labels = (np.where(predictions < 0.5, 0, 1)).flatten()
    return labels, model


def write_logs(fname, score, openpyxl=None):
    files = glob.glob(fname)
    fields = [score]
    export = False
    if len(files) != 0:
        sc_df = read_excel(fname)
        old = float(sc_df['Score'].values.tolist()[-1])
        if score > old:
            export = True
            from openpyxl import load_workbook
            # load the workbook, and put the sheet into a variable
            wb = load_workbook(filename=fname)
            ws = wb['Sheet1']

            # max_row is a sheet function that gets the last row in a sheet.
            newRowLocation = ws.max_row + 1
            # write to the cell you want, specifying row and column, and value :-)
            ws.cell(column=1, row=newRowLocation, value=score)
            wb.save(filename=fname)
            wb.close()
    return export


def export(classifier):
    classifier.save("my_model")
    print('keras model pickled')


if __name__ == '__main__':
    df = import_data(fname='datasets/cs-training.csv')
    print(list(df.columns))
    test_df = import_data(fname='datasets/cs-test.csv', train=False)
    df = clean_data(df)
    test_df = clean_data(test_df)
    df = normalize_columns(df, colnames=['age', 'MonthlyIncome', 'NumberOfDependents'])
    df = one_hot_encode(df, test_df, colnames=['ages'])
    df = undersample(df)
    X, Y, x_train, x_test, y_train, y_test = split_dataset(df, test_size=0.2, seed=42)
    missing = (X.isnull().values.any())
    if missing:
        print(X.isnull().sum())
    X_train, Y_train = np.array(x_train), np.array(y_train)
    classifier = get_model(X.shape[1], 1, magic='sigmoid')
    start = datetime.now()
    batch = 1024
    epochs = 5  # 100

    test_acc, test_loss = fit_and_evaluate(classifier, X_train, Y_train, x_test, y_test, batch_size=batch,
                                           epochs=epochs)
    end = datetime.now()
    total_seconds = (end - start).seconds
    minute, seconds = divmod(total_seconds, 60)
    print(f'\nTraining Time: {minute}:{seconds}')
    avg = total_seconds / epochs
    print(f'\nAverage time per epoch: {avg}')

    print(test_acc, test_loss)

    y_hat, model = make_predictions(classifier, x_test)

    roc_auc_score = get_metrics(y_test, y_hat, 'roc_auc_score', 'macro')
    print(roc_auc_score)

    fname = 'nn_logs.xlsx'
    export_flag = write_logs(fname, roc_auc_score)
    export(model)

    if export_flag:
        export(model)

    jcard_score = get_metrics(y_test, y_hat, 'jaccard_score', 'binary')
    f1 = get_metrics(y_test, y_hat, 'f1_score', 'binary')
    precision = get_metrics(y_test, y_hat, 'precision_score', 'binary')
    recall = get_metrics(y_test, y_hat, 'recall_score', 'binary')

    print('Program execution complete!')
